import os
import json
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
try:
    import win32com.client as win32
    import pythoncom
except Exception:
    win32 = None
    pythoncom = None

class ExcelManager:
    def __init__(self, filename="clinic_data.xlsx"):
        self.filename = filename
        self.setup_excel_file()
        
    def setup_excel_file(self):
        """Создание или загрузка Excel файла с нужными листами"""
        if os.path.exists(self.filename):
            # Загружаем существующий файл
            self.workbook = load_workbook(self.filename)
        else:
            # Создаем новый файл
            self.workbook = Workbook()
            # Удаляем дефолтный лист
            self.workbook.remove(self.workbook.active)
        
        # Создаем нужные листы если их нет
        self.create_sheet_if_not_exists('Записи на прием')
        self.create_sheet_if_not_exists('Отзывы')
        self.create_sheet_if_not_exists('Онлайн консультации')
        self.create_sheet_if_not_exists('Подписчики')
        
        # Устанавливаем заголовки
        self.setup_headers()
        
        # Сохраняем файл
        self.workbook.save(self.filename)
        
    def create_sheet_if_not_exists(self, sheet_name):
        """Создание листа если он не существует"""
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
            
    def setup_headers(self):
        """Установка заголовков для всех листов"""
        # Заголовки для записей на прием
        appointments_headers = [
            'Дата записи', 'Время', 'ФИО пациента', 'Телефон', 
            'Врач', 'Специализация', 'Статус', 'ID пользователя', 'Дата создания'
        ]
        
        # Заголовки для отзывов
        reviews_headers = [
            'Дата', 'ФИО', 'Оценка', 'Отзыв', 'ID пользователя', 'Статус'
        ]
        
        # Заголовки для онлайн консультаций
        consultations_headers = [
            'Дата', 'Вопрос', 'ID пользователя', 'Статус', 'Ответ'
        ]
        
        # Заголовки для подписчиков
        subscribers_headers = [
            'ID пользователя', 'Имя', 'Дата подписки'
        ]
        
        # Устанавливаем заголовки для каждого листа
        self.set_headers_for_sheet('Записи на прием', appointments_headers)
        self.set_headers_for_sheet('Отзывы', reviews_headers)
        self.set_headers_for_sheet('Онлайн консультации', consultations_headers)
        self.set_headers_for_sheet('Подписчики', subscribers_headers)
        
    def set_headers_for_sheet(self, sheet_name, headers):
        """Установка заголовков для конкретного листа"""
        sheet = self.workbook[sheet_name]
        
        # Проверяем, есть ли уже заголовки
        if sheet.max_row == 0 or sheet.cell(row=1, column=1).value is None:
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        # Сохраняем изменения
        self.workbook.save(self.filename)

    # ===== COM-помощники для работы с открытым Excel =====
    def _com_open_workbook(self):
        """Открыть/найти книгу в Excel через COM. Возвращает (excel, workbook, opened_here)."""
        if win32 is None or pythoncom is None:
            return None, None, False
        try:
            pythoncom.CoInitialize()
            try:
                excel = win32.Dispatch("Excel.Application")
            except Exception:
                excel = win32.DispatchEx("Excel.Application")
            excel.DisplayAlerts = False
            target = os.path.abspath(self.filename)
            wb = None
            for w in excel.Workbooks:
                try:
                    if os.path.abspath(str(w.FullName)) == target:
                        wb = w
                        break
                except Exception:
                    continue
            opened_here = False
            if wb is None:
                try:
                    wb = excel.Workbooks.Open(target, UpdateLinks=0, ReadOnly=False)
                    opened_here = True
                except Exception:
                    return None, None, False
            return excel, wb, opened_here
        except Exception:
            return None, None, False

    def _com_finalize(self, excel, wb, opened_here):
        try:
            if wb is not None:
                try:
                    wb.Save()
                except Exception:
                    pass
                if opened_here:
                    try:
                        wb.Close(SaveChanges=True)
                    except Exception:
                        pass
            if excel is not None:
                try:
                    excel.DisplayAlerts = True
                except Exception:
                    pass
        finally:
            try:
                if pythoncom is not None:
                    pythoncom.CoUninitialize()
            except Exception:
                pass

    def _com_append_row(self, sheet_name, values):
        """Добавить строку в конец указанного листа через COM. Возвращает True/False."""
        excel, wb, opened_here = self._com_open_workbook()
        if wb is None:
            return False
        try:
            try:
                ws = wb.Worksheets(sheet_name)
            except Exception:
                ws = wb.Worksheets.Add()
                ws.Name = sheet_name
            # Найти последнюю заполненную строку
            try:
                last_row = ws.Cells(ws.Rows.Count, 1).End(-4162).Row  # xlUp
                if last_row < 1:
                    last_row = 1
            except Exception:
                last_row = 1
            next_row = last_row + 1
            for idx, val in enumerate(values, start=1):
                try:
                    ws.Cells(next_row, idx).Value = val
                except Exception:
                    ws.Cells(next_row, idx).Value = str(val) if val is not None else ""
            self._com_finalize(excel, wb, opened_here)
            return True
        except Exception:
            self._com_finalize(excel, wb, opened_here)
            return False

    def _com_delete_appointment_row(self, user_id, date, time, doctor, created_at):
        """Удалить строку в листе 'Записи на прием' по ключам через COM."""
        excel, wb, opened_here = self._com_open_workbook()
        if wb is None:
            return False
        try:
            try:
                ws = wb.Worksheets('Записи на прием')
            except Exception:
                self._com_finalize(excel, wb, opened_here)
                return False
            try:
                used_rows = ws.UsedRange.Rows.Count
            except Exception:
                used_rows = 0
            target_user = str(user_id)
            target_date = str(date)
            target_time = str(time)
            target_doctor = str(doctor)
            target_created = str(created_at)
            for r in range(2, used_rows + 1):
                try:
                    v_user = str(ws.Cells(r, 8).Value)
                    v_date = ws.Cells(r, 1).Value
                    try:
                        v_date = v_date.strftime('%d.%m.%Y') if hasattr(v_date, 'strftime') else str(v_date)
                    except Exception:
                        v_date = str(v_date)
                    v_time = str(ws.Cells(r, 2).Value)
                    v_doctor = str(ws.Cells(r, 5).Value)
                    v_created = str(ws.Cells(r, 9).Value)
                except Exception:
                    continue
                if (
                    v_user == target_user and
                    v_date == target_date and
                    v_time == target_time and
                    v_doctor == target_doctor and
                    v_created == target_created
                ):
                    try:
                        ws.Rows(r).Delete()
                        self._com_finalize(excel, wb, opened_here)
                        return True
                    except Exception:
                        break
            self._com_finalize(excel, wb, opened_here)
            return False
        except Exception:
            self._com_finalize(excel, wb, opened_here)
            return False

    # Совместимость: фоновой сброс (нет очереди) — делаем no-op
    def flush_pending_ops(self):
        return False
        
    def add_appointment(self, date, time, patient_name, phone, doctor, specialization, user_id):
        """Добавление записи на прием"""
        try:
            # Переключаемся на pandas для защиты от дублей
            try:
                df = pd.read_excel(self.filename, sheet_name='Записи на прием')
            except Exception:
                df = pd.DataFrame(columns=[
                    'Дата записи','Время','ФИО пациента','Телефон','Врач','Специализация','Статус','ID пользователя','Дата создания'
                ])

            # Удаляем точные дубли по пользователю/доктору/дата/время, оставляем последнюю
            df = df.sort_values('Дата создания').drop_duplicates(
                subset=['ID пользователя','Дата записи','Время','Врач'], keep='last'
            )

            new_row = {
                'Дата записи': date,
                'Время': time,
                'ФИО пациента': patient_name,
                'Телефон': phone,
                'Врач': doctor,
                'Специализация': specialization,
                'Статус': 'Новая',
                'ID пользователя': str(user_id),
                'Дата создания': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

            # Сохраняем обратно в Excel поверх существующего файла
            try:
                with pd.ExcelWriter(self.filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    # Полная замена листа, чтобы не оставались старые строки
                    df.to_excel(writer, sheet_name='Записи на прием', index=False)
                print('Запись добавлена (pandas)')
                return True
            except Exception as e:
                # Попробуем напрямую через COM (если Excel открыт)
                if win32 is not None:
                    values = [
                        date,
                        time,
                        patient_name,
                        phone,
                        doctor,
                        specialization,
                        'Новая',
                        str(user_id),
                        new_row['Дата создания']
                    ]
                    from_types = [str, str, str, str, str, str, str, str, str]
                    # Преобразуем значения к строке для надёжной записи через COM
                    values = [str(v) if v is not None else '' for v in values]
                    if self._com_append_row('Записи на прием', values):
                        print('Запись добавлена через COM')
                        return True
                print(f'Ошибка записи в файл: {e}')
                return False
            
        except Exception as error:
            print(f'Ошибка при добавлении записи: {error}')
            return False
    
    def add_review(self, patient_name, rating, review_text, user_id):
        """Добавление отзыва"""
        try:
            sheet = self.workbook['Отзывы']
            
            # Находим следующую пустую строку
            next_row = sheet.max_row + 1
            
            # Добавляем данные
            data = [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                patient_name,
                rating,
                review_text,
                str(user_id),
                'Новый'
            ]
            
            for col, value in enumerate(data, 1):
                sheet.cell(row=next_row, column=col, value=value)
            
            # Сохраняем файл
            try:
                self.workbook.save(self.filename)
                print(f'Отзыв добавлен в строку {next_row}')
                return True
            except Exception as e:
                # Пробуем COM
                if win32 is not None:
                    values = [
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        patient_name,
                        rating,
                        review_text,
                        str(user_id),
                        'Новый'
                    ]
                    values = [str(v) if v is not None else '' for v in values]
                    if self._com_append_row('Отзывы', values):
                        print('Отзыв добавлен через COM')
                        return True
                print(f'Ошибка записи отзыва: {e}')
                return False
            
        except Exception as error:
            print(f'Ошибка при добавлении отзыва: {error}')
            return False
    
    def add_consultation(self, question, user_id):
        """Добавление онлайн консультации"""
        try:
            sheet = self.workbook['Онлайн консультации']
            
            # Находим следующую пустую строку
            next_row = sheet.max_row + 1
            
            # Добавляем данные
            data = [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                question,
                str(user_id),
                'Новая',
                ''
            ]
            
            for col, value in enumerate(data, 1):
                sheet.cell(row=next_row, column=col, value=value)
            
            # Сохраняем файл
            try:
                self.workbook.save(self.filename)
                print(f'Консультация добавлена в строку {next_row}')
                return True
            except Exception as e:
                if win32 is not None:
                    values = [
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        question,
                        str(user_id),
                        'Новая',
                        ''
                    ]
                    values = [str(v) if v is not None else '' for v in values]
                    if self._com_append_row('Онлайн консультации', values):
                        print('Консультация добавлена через COM')
                        return True
                print(f'Ошибка записи консультации: {e}')
                return False
            
        except Exception as error:
            print(f'Ошибка при добавлении консультации: {error}')
            return False
    
    def add_subscriber(self, user_id, user_name):
        """Добавление подписчика"""
        try:
            sheet = self.workbook['Подписчики']
            
            # Проверяем, не подписан ли уже пользователь
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row=row, column=1).value == str(user_id):
                    print(f'Пользователь {user_id} уже подписан')
                    return True
            
            # Находим следующую пустую строку
            next_row = sheet.max_row + 1
            
            # Добавляем данные
            data = [
                str(user_id),
                user_name,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            for col, value in enumerate(data, 1):
                sheet.cell(row=next_row, column=col, value=value)
            
            # Сохраняем файл
            try:
                self.workbook.save(self.filename)
                print(f'Подписчик добавлен в строку {next_row}')
                return True
            except Exception as e:
                if win32 is not None:
                    values = [
                        str(user_id),
                        user_name,
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                    values = [str(v) if v is not None else '' for v in values]
                    if self._com_append_row('Подписчики', values):
                        print('Подписчик добавлен через COM')
                        return True
                print(f'Ошибка записи подписчика: {e}')
                return False
            
        except Exception as error:
            print(f'Ошибка при добавлении подписчика: {error}')
            return False
    
    def get_appointments(self):
        """Получение всех записей на прием"""
        try:
            df = pd.read_excel(self.filename, sheet_name='Записи на прием')
            if df.empty:
                return []
            
            # Конвертируем DataFrame в список списков
            return df.values.tolist()
            
        except Exception as error:
            print(f'Ошибка при получении записей: {error}')
            return []
    
    def get_reviews(self):
        """Получение всех отзывов"""
        try:
            df = pd.read_excel(self.filename, sheet_name='Отзывы')
            if df.empty:
                return []
            
            # Конвертируем DataFrame в список списков
            return df.values.tolist()
            
        except Exception as error:
            print(f'Ошибка при получении отзывов: {error}')
            return []
    
    def get_subscribers(self):
        """Получение всех подписчиков"""
        try:
            df = pd.read_excel(self.filename, sheet_name='Подписчики')
            if df.empty:
                return []
            
            # Конвертируем DataFrame в список списков
            return df.values.tolist()
            
        except Exception as error:
            print(f'Ошибка при получении подписчиков: {error}')
            return []
    
    def get_consultations(self):
        """Получение всех консультаций"""
        try:
            df = pd.read_excel(self.filename, sheet_name='Онлайн консультации')
            if df.empty:
                return []
            
            # Конвертируем DataFrame в список списков
            return df.values.tolist()
            
        except Exception as error:
            print(f'Ошибка при получении консультаций: {error}')
            return []
    
    def update_appointment_status(self, row_index, new_status):
        """Обновление статуса записи на прием"""
        try:
            sheet = self.workbook['Записи на прием']
            sheet.cell(row=row_index, column=7, value=new_status)  # Колонка G - статус
            self.workbook.save(self.filename)
            return True
        except Exception as error:
            print(f'Ошибка при обновлении статуса: {error}')
            return False
    
    def update_review_status(self, row_index, new_status):
        """Обновление статуса отзыва"""
        try:
            sheet = self.workbook['Отзывы']
            sheet.cell(row=row_index, column=6, value=new_status)  # Колонка F - статус
            self.workbook.save(self.filename)
            return True
        except Exception as error:
            print(f'Ошибка при обновлении статуса отзыва: {error}')
            return False
    
    def backup_data(self, backup_filename=None):
        """Создание резервной копии данных"""
        if backup_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            backup_filename = f"backup_clinic_data_{timestamp}.xlsx"
        
        try:
            self.workbook.save(backup_filename)
            print(f'Резервная копия создана: {backup_filename}')
            return True
        except Exception as error:
            print(f'Ошибка при создании резервной копии: {error}')
            return False

    def get_appointments_by_user(self, user_id):
        """Получение всех записей на прием для конкретного пользователя (чтение напрямую из файла через pandas)."""
        try:
            df = pd.read_excel(self.filename, sheet_name='Записи на прием')
            if df.empty:
                return []
            # Приводим колонку ID к строке для сопоставления
            df = df[df['ID пользователя'].astype(str) == str(user_id)]
            if df.empty:
                return []
            # Удаляем потенциальные дубликаты: оставляем самую позднюю запись по 'Дата создания'
            df = df.sort_values('Дата создания').drop_duplicates(subset=['Дата записи', 'Время', 'Врач'], keep='last')
            # Возвращаем в требуемом порядке столбцы
            cols = ['Дата записи', 'Время', 'ФИО пациента', 'Телефон', 'Врач', 'Специализация', 'Статус', 'ID пользователя', 'Дата создания']
            df = df[cols]
            return df.values.tolist()
        except Exception as error:
            print(f'Ошибка при получении записей пользователя: {error}')
            return []

    def get_booked_times(self, doctor_name, date):
        """Получить занятые слоты времени для врача на конкретную дату (чтение напрямую из файла)."""
        try:
            df = pd.read_excel(self.filename, sheet_name='Записи на прием')
            if df.empty:
                return set()
            # Нормализуем дату записи в строку dd.mm.YYYY для сравнения
            def norm_date(val):
                try:
                    return val.strftime('%d.%m.%Y')
                except Exception:
                    return str(val)
            df['__norm_date'] = df['Дата записи'].apply(norm_date)
            # Удаляем дубликаты по слоту (берём последнюю по дате создания)
            df = df.sort_values('Дата создания').drop_duplicates(subset=['Дата записи', 'Время', 'Врач'], keep='last')

            mask = (
                df['Врач'].astype(str) == str(doctor_name)
            ) & (
                df['__norm_date'].astype(str) == str(date)
            ) & (
                ~df['Статус'].astype(str).str.lower().isin(['отменена', 'отменён', 'cancelled', 'canceled'])
            )
            booked = set(df.loc[mask, 'Время'].astype(str).tolist())
            return booked
        except Exception as error:
            print(f'Ошибка при получении занятых слотов: {error}')
            return set()

    def delete_appointment(self, user_id, date, time, doctor, created_at):
        """Удалить запись на прием по ключевым полям (пользователь, дата, время, врач, создано) через pandas."""
        try:
            df = pd.read_excel(self.filename, sheet_name='Записи на прием')
            if df.empty:
                return False

            def normalize_date(value):
                try:
                    return value.strftime('%d.%m.%Y')
                except Exception:
                    return str(value)

            def normalize_created_at(value):
                try:
                    return value.strftime('%Y-%m-%d %H:%M:%S')
                except Exception:
                    return str(value)

            df['__norm_date'] = df['Дата записи'].apply(normalize_date)
            df['__norm_time'] = df['Время'].astype(str)
            df['__norm_doctor'] = df['Врач'].astype(str)
            df['__norm_user'] = df['ID пользователя'].astype(str)
            df['__norm_created'] = df['Дата создания'].apply(normalize_created_at)

            target_date = str(date)
            target_time = str(time)
            target_doctor = str(doctor)
            target_user = str(user_id)
            target_created = str(created_at)

            before = len(df)
            mask = (
                (df['__norm_user'] == target_user) &
                (df['__norm_date'] == target_date) &
                (df['__norm_time'] == target_time) &
                (df['__norm_doctor'] == target_doctor) &
                (df['__norm_created'] == target_created)
            )
            df = df.loc[~mask, :]

            if len(df) == before:
                return False

            # Очистим служебные столбцы
            df = df.drop(columns=[c for c in df.columns if c.startswith('__norm_')], errors='ignore')

            # Полная замена листа
            try:
                with pd.ExcelWriter(self.filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df.to_excel(writer, sheet_name='Записи на прием', index=False)
                print(f'Удалена запись пользователя {user_id} на {date} {time} к {doctor}')
                return True
            except Exception as e:
                # Попытка через COM: найдём и удалим строку
                if win32 is not None:
                    if self._com_delete_appointment_row(user_id, date, time, doctor, created_at):
                        print('Удаление записи выполнено через COM')
                        return True
                print(f'Ошибка при сохранении удаления: {e}')
                return False
        except Exception as error:
            print(f'Ошибка при удалении записи: {error}')
            return False